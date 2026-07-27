"""
Builds docs/dashboard.xlsx - one Excel workbook combining every category's
search-trend history, formatted to match the original Babycare-style
analysis sheet: bold headers, Combined Index / Trailing Average / Rolling
Growth % columns, and an embedded dark-themed line chart.
"""
import csv
import glob
import os
import re

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import (
    CharacterProperties, Paragraph, ParagraphProperties, RichTextProperties,
)
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import ColorChoice
from openpyxl.styles import Font

DATA_DIR = "data"
TRAILING_PERIODS = 8  # rolling window, in weeks, for the moving average


def slug_to_label(path):
    name = os.path.basename(path)
    name = re.sub(r"^trends_", "", name)
    return re.sub(r"\.csv$", "", name).replace("-", " ")


def load_csv(path):
    with open(path) as f:
        return list(csv.DictReader(f))


def main():
    os.makedirs("docs", exist_ok=True)
    files = sorted(glob.glob(f"{DATA_DIR}/trends_*.csv"))
    if not files:
        print("No data yet - skipping workbook export")
        return

    series = {}       # category -> {date: index}
    all_dates = set()
    for path in files:
        rows = load_csv(path)
        label = rows[0]["category"] if rows else slug_to_label(path)
        by_date = {r["date"]: float(r["search_index"]) for r in rows if r.get("search_index")}
        series[label] = by_date
        all_dates.update(by_date.keys())

    dates = sorted(all_dates)
    categories = sorted(series.keys())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Search Trends Analysis"

    headers = ["Date"] + categories + ["Combined Index", "Trailing Average", "Rolling Growth %"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    n_cat = len(categories)
    for i, d in enumerate(dates):
        row_idx = i + 2
        row = [d] + [series[c].get(d, "") for c in categories]
        ws.append(row)

        first_col = openpyxl.utils.get_column_letter(2)
        last_col = openpyxl.utils.get_column_letter(1 + n_cat)
        combined_col = openpyxl.utils.get_column_letter(2 + n_cat)
        avg_col = openpyxl.utils.get_column_letter(3 + n_cat)
        growth_col = openpyxl.utils.get_column_letter(4 + n_cat)

        ws[f"{combined_col}{row_idx}"] = f"=SUM({first_col}{row_idx}:{last_col}{row_idx})"

        if i >= TRAILING_PERIODS - 1:
            start = row_idx - TRAILING_PERIODS + 1
            ws[f"{avg_col}{row_idx}"] = f"=AVERAGE({combined_col}{start}:{combined_col}{row_idx})"
        if i >= TRAILING_PERIODS:
            prev_row = row_idx - TRAILING_PERIODS
            ws[f"{growth_col}{row_idx}"] = (
                f"=IF({avg_col}{prev_row}=0,\"\",({avg_col}{row_idx}-{avg_col}{prev_row})/{avg_col}{prev_row})"
            )
            ws[f"{growth_col}{row_idx}"].number_format = "0.0%"

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16

    # --- chart: dark-themed line chart of Rolling Growth %, matching the
    #     original workbook's style (bold all-caps title, dark background) ---
    if len(dates) > TRAILING_PERIODS:
        chart = LineChart()
        chart.style = 26  # built-in dark chart style
        chart.height = 9
        chart.width = 20

        growth_col_idx = 4 + n_cat
        data_ref = Reference(ws, min_col=growth_col_idx, min_row=1, max_row=1 + len(dates))
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=1 + len(dates))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        chart.title = "SEARCHES YoY GROWTH TREND"
        title_para = chart.title.tx.rich.p[0]
        title_para.pPr = ParagraphProperties(
            defRPr=CharacterProperties(sz=1500, b=True, cap="all",
                                        solidFill=ColorChoice(srgbClr="FFFFFF"))
        )
        chart.graphical_properties = GraphicalProperties(solidFill="1F1F1F")

        ws.add_chart(chart, f"A{len(dates) + 4}")

    wb.save("docs/dashboard.xlsx")
    print(f"Workbook written to docs/dashboard.xlsx ({len(dates)} weeks x {n_cat} categories)")


if __name__ == "__main__":
    main()
